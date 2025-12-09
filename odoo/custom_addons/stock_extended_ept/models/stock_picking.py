# -*- coding: utf-8 -*-

import io
import base64
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from datetime import datetime

from odoo import models, fields, _
from odoo.exceptions import ValidationError


class StockPickingExtend(models.Model):
    _inherit = 'stock.picking'

    bayan_code = fields.Char(string="Bayan Code")
    is_bayan_code_visible = fields.Boolean(
        related='company_id.is_bayan_code_applicable',
        store=False
    )
    boe_number = fields.Char(string="BOE Number", help="Bill Of Entry Number")

    def button_validate(self):
        """
        Writes bayan_code to related lots when validating incoming pickings
        """
        res = super().button_validate()
        for picking in self:
            if picking.picking_type_id.code == 'incoming':
                # Incoming picking logic
                if picking.bayan_code:
                    lot_recs = picking.move_line_ids.mapped('lot_id')
                    for lot in lot_recs:
                        lot.write({
                            'bayan_code': picking.bayan_code,
                        })

                if not picking.bayan_code or not picking.boe_number:
                    missing_fields = []
                    if picking.company_id.is_bayan_code_applicable and not picking.bayan_code:
                        missing_fields.append("Bayan Code")
                    if not picking.boe_number:
                        missing_fields.append("BOE Number")
                    if missing_fields:
                        raise ValidationError(_(
                            "The following fields are missing and are mandatory to validate the receipt:\n%s",
                            '\n'.join(f" - {field}" for field in missing_fields)
                        ))
            elif picking.picking_type_id.code == 'outgoing':
                # Outgoing picking logic - only BOE required
                if not picking.boe_number:
                    raise ValidationError(_(
                        "The following fields are missing and are mandatory to validate the receipt: \n - Bill Of Entry Number"
                    ))
        return res

    def action_download_packing_slip_xlsx_report(self):
        """
        Prepares packing slip data, creates xlsx attachment and returns a download action
        """
        data = self.prepare_packing_slip_report_data()
        attachment = self.create_xlsx_data_sheet(data)
        # Return download action
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/{attachment.id}?download=true",
            'target': 'self'
        }

    def prepare_packing_slip_report_data(self):
        """
        Gathers delivery order and product details into a dict for xlsx export
        """
        picking_type = self.picking_type_id.code
        shipping_date = self.date_done if self.state == 'done' else self.scheduled_date
        shipping_weight = self.shipping_weight or 0.0
        do_data = {'Transfer': self.display_name}

        if picking_type == 'incoming':
            do_data.update(self._get_incoming_data())
        else:
            do_data.update(self._get_outgoing_data())

        do_data.update(self._get_common_data(shipping_date, shipping_weight))
        do_data.update(self._get_package_data())
        do_data.update(self._get_backorder_data())

        return do_data

    def _get_incoming_data(self):
        """Get incoming order specific data"""
        incoterm = self.purchase_id.incoterm_id
        incoterm_location = self.purchase_id.incoterm_location
        invoices = self.move_ids.mapped('purchase_line_id.invoice_lines.move_id.name')

        return {
            'Vendor': self.partner_id.name,
            'Order': self.purchase_id.name or 'N/A',
            'Vendor Reference': self.purchase_id.partner_ref or 'N/A',
            'Incoterm': "'%s %s' % (incoterm.code, incoterm_location)" if (
                incoterm_location) else incoterm.display_name or 'N/A',
            'Invoice Number': ", ".join(invoices) if invoices else 'N/A',
        }

    def _get_outgoing_data(self):
        """Get outgoing order specific data"""
        incoterm = self.sale_id.incoterm
        incoterm_location = self.sale_id.incoterm_location
        invoices = self.move_ids.mapped('sale_line_id.invoice_lines.move_id.name')

        return {
            'Customer': self.partner_id.name,
            'Order': self.sale_id.name or 'N/A',
            'Customer Reference': self.sale_id.client_order_ref or 'N/A',
            'Incoterm': "'%s %s' % (incoterm.code, incoterm_location)" if (
                incoterm_location) else incoterm.display_name or 'N/A',
            'Invoice Number': ", ".join(invoices) if invoices else 'N/A',
        }

    def _get_common_data(self, shipping_date, shipping_weight):
        """Get common data for both incoming and outgoing"""
        common_data = {
            'Shipping Date': shipping_date.strftime('%Y-%m-%d') if shipping_date else 'N/A',
            'Total Weight': f"{shipping_weight} {self.weight_uom_name}",
            'Tracking Number': self.carrier_tracking_ref or 'N/A',
        }
        if self.bayan_code:
            common_data.update({'Bayan Code': self.bayan_code or 'N/A'})
        if self.boe_number:
            common_data.update({'BOE Number': self.boe_number or 'N/A'})
        return common_data


    def _get_package_data(self):
        """Get package and without package details"""
        package_data = {}

        if self.has_packages:
            package_data.update(self._get_package_details())
            package_data.update(self._get_without_package_details_for_has_packages())
        else:
            package_data.update(self._get_without_package_details_for_no_packages())

        return package_data

    def _get_package_details(self):
        """Get package details when has_packages is True"""
        package_details = {}

        for pack in self.move_line_ids.mapped('result_package_id'):
            move_lines = self.move_line_ids.filtered(lambda l: l.result_package_id == pack)
            for line in move_lines:
                if not package_details.get('Package Details'):
                    package_details['Package Details'] = {}
                if pack not in package_details.get('Package Details', {}):
                    package_details['Package Details'].update({
                        pack: {
                            line.id: self._get_line_data(line)
                        }
                    })
                else:
                    package_details['Package Details'][pack].update({
                        line.id: self._get_line_data(line)
                    })

        return package_details

    def _get_without_package_details_for_has_packages(self):
        """Get without package details when has_packages is True"""
        without_package_details = {}

        for line in self.move_line_ids.filtered(lambda l: not l.result_package_id):
            if not without_package_details.get('Without Package Details'):
                without_package_details['Without Package Details'] = {}
            without_package_details['Without Package Details'].update({
                line.id: self._get_line_data(line)
            })

        return without_package_details

    def _get_without_package_details_for_no_packages(self):
        """Get without package details when has_packages is False"""
        without_package_details = {}

        for move in self.move_ids.filtered(lambda x: x.product_uom_qty):
            for line in move.move_line_ids:
                if not without_package_details.get('Without Package Details'):
                    without_package_details['Without Package Details'] = {}
                without_package_details['Without Package Details'].update({
                    line.id: self._get_line_data(line)
                })
        return without_package_details

    def _get_line_data(self, line):
        """Get line data for package and non-package items"""
        # 'Internal Ref.': line.product_id.default_code or 'N/A',
        return {
            'Model Number': line.move_id.model_number or 'N/A',
            'Product': line.product_id.name,
            'Lot/Serial': line.lot_id.name or 'N/A',
            'Delivered': line.quantity,
            'Weight': f'{line.product_id.weight * line.quantity}  {self.weight_uom_name}'
        }

    def _get_backorder_data(self):
        backorder_data = {}
        back_orders = self.backorder_ids.filtered(lambda x: x.state not in ('done', 'cancel'))
        for move in back_orders.move_ids.filtered(lambda x: x.product_uom_qty):
            if not backorder_data.get('Backorder Details'):
                backorder_data['Backorder Details'] = {}
            backorder_data['Backorder Details'].update({
                # 'Internal Ref.': move.product_id.default_code or 'N/A',
                move.id: {
                    'Model Number': move.model_number or 'N/A',
                    'Product': move.product_id.name,
                    'Quantity': move.product_uom_qty,
                }
            })
        return backorder_data

    def create_xlsx_data_sheet(self, data):
        """
        Builds an openpyxl workbook from provided data and returns an ir.attachment record
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Delivery Details"

        self.set_column_dimensions(ws)

        do_data = {k: v for k, v in data.items() if k not in ['Package Details', 'Without Package Details', 'Backorder Details']}
        product_data = {k: v for k, v in data.items() if k in ['Package Details', 'Without Package Details']}
        backorder_data = data.get('Backorder Details', {})

        # Delivery Order Details
        row = 1
        for key, value in do_data.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Product Details
        if product_data:
            headers, has_lot = self.has_lot_serial_data(product_data)

            # Product Headers
            row += 2  # Adding some space before product details
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
            row += 1



            if product_data.get('Package Details'):
                pack_data = product_data['Package Details']
                row = self.add_package_details_in_slip(pack_data, headers, row, ws)
            if product_data.get('Without Package Details'):
                row = self.add_without_package_details_in_slip(product_data, headers, row, ws)

        # Backorder details
        if backorder_data:
            self.add_backorder_details_in_slip(backorder_data, headers, row, ws)
        return self.create_packing_slip_xlsx_attachment(wb)

    def add_package_details_in_slip(self, pack_data, headers, row, ws):
        """
        Adds package details to the packing slip worksheet
        """
        for pack, line_dict in pack_data.items():
            # Pack title
            row += 1
            pack_title = self.get_pack_title(pack)
            cell = ws.cell(row=row, column=1)
            cell.value = pack_title
            cell.font = Font(bold=True, size=9)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))

            row += 1
            for line_id, line_vals in line_dict.items():
                for j, col in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=j)
                    cell.value = line_vals.get(col, "")
                    cell.alignment = Alignment(horizontal='left')
                row += 1
        return row

    def add_backorder_details_in_slip(self, backorder_data, headers, row, ws):
        row += 3  # Adding some space before backorder details
        ws.cell(row=row, column=1, value="Remaining quantities not yet delivered:")
        ws.cell(row=row, column=1).font = Font(bold=True, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1

        # Backorder Headers
        first_key = next(iter(backorder_data))
        backorder_headers = list(backorder_data[first_key].keys())
        row += 1
        for col_num, header in enumerate(backorder_headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        row += 1

        for move, line_vals in backorder_data.items():
            for j, col in enumerate(backorder_headers, 1):
                cell = ws.cell(row=row, column=j)
                cell.value = line_vals.get(col, "")
                cell.alignment = Alignment(horizontal='left')
            row += 1

    def add_without_package_details_in_slip(self, product_data, headers, row, ws):
        """
        Adds without package details to the packing slip worksheet
        """
        if product_data.get('Package Details'):
            row += 1
            cell = ws.cell(row=row, column=1)
            cell.value = "Products with no package assigned"
            cell.font = Font(bold=True, size=9)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1
        line_dict = product_data.get('Without Package Details')
        for line_id, line_vals in line_dict.items():
            for j, col in enumerate(headers, 1):
                cell = ws.cell(row=row, column=j)
                cell.value = line_vals.get(col, "")
                cell.alignment = Alignment(horizontal='left')
            row += 1
        return row

    def get_pack_title(self, pack):
        title = pack.name
        if pack.weight:
            title += f" - {pack.weight} {pack.weight_uom_name}"
        if any([pack.package_type_id.packaging_length, pack.package_type_id.width,
                pack.package_type_id.height]):
            title += (f" - {pack.package_type_id.packaging_length or 0} x "
                      f"{pack.package_type_id.width or 0} x {pack.package_type_id.height or 0} "
                      f"{pack.package_type_id.length_uom_name}")
        return title

    def has_lot_serial_data(self, data):
        has_lot = False
        headers = []
        for package, products in data.get('Package Details', {}).items():
            for line, line_data in products.items():
                if not headers:
                    headers = list(line_data.keys())
                if line_data.get('Lot/Serial') not in ['N/A', '', None]:
                    has_lot = True
                    return headers, has_lot

        # Check Without Package Details
        for line, line_data in data.get('Without Package Details', {}).items():
            if not headers:
                headers = list(line_data.keys())
            if line_data.get('Lot/Serial') not in ['N/A', '', None]:
                has_lot = True
                return headers, has_lot
        headers = [h for h in headers if h != 'Lot/Serial']
        return headers, has_lot

    @staticmethod
    def set_column_dimensions(ws):
        """
        Sets column widths for the packing slip worksheet
        """
        widths = {
            'A': 18, 'B': 40, 'C': 15, 'D': 15
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def create_packing_slip_xlsx_attachment(self, wb):
        """
        Saves workbook to bytes and creates an ir.attachment for download
        """
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        file_content = stream.getvalue()
        fmt = '%m-%d-%Y'

        return self.env['ir.attachment'].create({
            'name': f'Packing Slip({datetime.now().strftime(fmt)}).xlsx',
            'type': 'binary',
            'datas': base64.b64encode(file_content),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
