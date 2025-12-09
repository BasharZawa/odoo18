# -*- coding: utf-8 -*-

import io
import base64
from datetime import datetime
from bs4 import BeautifulSoup
from odoo import models, fields
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ProductVariantExtended(models.Model):
    """
    Add logic related to details related to current stock of product
    """
    _inherit = 'product.product'

    qty_available_threshold = fields.Float(string="Low Quantity Threshold")
    tolerance = fields.Float(string="Allowed Tolerance (%)", digits=(4, 2))

    def action_generate_pending_report(self):
        """
        Server action to call up method to generate current stock details record
        :return: dictionary
        """
        stock_data = self.prepare_current_stock_data()
        if stock_data:
            attachment = self.create_xlsx_attachment(stock_data)
            return {
                'type': 'ir.actions.act_url',
                'url': f"/web/content/{attachment.id}?download=true",
                'target': 'self'
            }

    def prepare_current_stock_data(self):
        """
        Prepare stock details for product variants
        :return: stock_data - List of dictionary
        """
        stock_data = []
        for prod in self.env['product.product'].search([('type', '=', 'consu')]):
            product_so_lines = self.get_prod_sale_order_lines(prod)
            product_po_lines = self.get_prod_purchase_order_lines(prod)
            total_sales = sum([sol.product_uom_qty - sol.qty_delivered for sol in product_so_lines])
            po_qty = sum([pol.product_qty - pol.qty_received for pol in product_po_lines])
            required_po_qty = total_sales - prod.qty_available - po_qty
            current_net_before = total_sales - prod.qty_available
            vendors = ', '.join(prod.seller_ids.mapped('partner_id.name'))
            supplier_pos = supplier_eta = ''
            fmt = '%m/%d/%Y'
            if product_po_lines:
                supplier_pos = ', '.join(product_po_lines.mapped('order_id.name'))
                supplier_eta = ', '.join(
                    [date.strftime(fmt)
                     for date in product_po_lines.mapped('order_id.date_planned')])
            data = {
                'Item Code': prod.default_code,
                'Item Description': prod.name or '',
                'Current Cost': prod.standard_price,
                'Short Description': prod.description or '',
                'Supplier': vendors,
                'Total Sales\nOrder Qty': total_sales,
                'Current Stock': prod.qty_available,
                'On P.Order Qty': po_qty,
                'Required PO\nQty': required_po_qty,
                'Current Net\nBefore': current_net_before,
                'Supplier PO No': supplier_pos,
                'ETA': supplier_eta
            }
            stock_data.append(data)
        return stock_data

    def create_xlsx_attachment(self, stock_data):
        """
        Creates XLSX attachment for stock data.
        :param stock_data: Prepared stock details to include in report
        :return: New attachment record
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Data"

        self.stock_report_set_column_widths(ws)
        self.stock_report_add_headers(ws, stock_data)
        self.stock_report_add_rows(ws, stock_data)

        return self.stock_report_save_xlsx_to_attachment(wb)

    @staticmethod
    def stock_report_set_column_widths(ws):
        """
        Adjusts column widths for the worksheet.
        :param ws: Worksheet object
        :return: None
        """
        widths = {
            'A': 15, 'B': 30, 'C': 13, 'D': 20, 'E': 20,
            'F': 13, 'G': 15, 'H': 15, 'I': 15, 'J': 15,
            'K': 20, 'L': 15
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for col in range(13, 30):
            ws.column_dimensions[get_column_letter(col)].width = 15

    @staticmethod
    def stock_report_add_headers(ws, stock_data):
        """
        Adds header row to worksheet.
        :param ws: Worksheet object
        :param stock_data: Stock data dictionary
        :return: None
        """
        headers = list(stock_data[0].keys())
        ws.append(headers)
        ws.row_dimensions[1].height = 30

        # Style headers
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)

    def stock_report_add_rows(self, ws, stock_data):
        """
        Adds data rows to worksheet.
        :param ws: Worksheet object
        :param stock_data: Stock data dictionary
        :return: None
        """
        headers = list(stock_data[0].keys())
        for row in stock_data:
            row_data = []
            for col in headers:
                if col == 'Short Description':
                    html_content = row.get(col, "")
                    text_content = BeautifulSoup(html_content, "html.parser").get_text()
                    row_data.append(text_content)
                else:
                    row_data.append(row.get(col, ""))
            if row.get('Item Code', ''):
                self.add_so_qty_to_xlsx(row.get('Item Code', ''), row_data)
            ws.append(row_data)

    def stock_report_save_xlsx_to_attachment(self, wb):
        """
        Saves workbook in memory and creates an Odoo attachment.
        :param wb: Workbook object
        :return: New attachment record
        """
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        file_content = stream.getvalue()

        fmt = '%m-%d-%Y'
        attachment_name = f"Self Service Components availability({datetime.now().strftime(fmt)}).xlsx"

        return self.env['ir.attachment'].create({
            'name': attachment_name,
            'type': 'binary',
            'datas': base64.b64encode(file_content),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

    def add_so_qty_to_xlsx(self, sku, row_data):
        """
        Prepare extra column data to report for sale order details
        :param sku: SKU of product
        :param row_data: prepared detail for report
        :return: None
        """
        prod = self.env['product.product'].search([('default_code', '=', sku)])
        product_so_lines = self.get_prod_sale_order_lines(prod)
        data = {}
        for sol in product_so_lines:
            if sol.order_id.name in data:
                data[sol.order_id.name] += sol.product_uom_qty - sol.qty_delivered
            else:
                data[sol.order_id.name] = sol.product_uom_qty - sol.qty_delivered
        for so, qty in data.items():
            row_data.append(f'{so} -> {qty}')

    def get_prod_sale_order_lines(self, prod):
        """
        Returns sale order line for given product
        :param prod: product object
        :return: Objects of sale order line for that product
        """
        s_order_lines = self.env['sale.order.line'].search(
            [('product_id', '=', prod.id), ('order_id.state', '!=', 'cancel'),
             ('order_id.delivery_status', '!=', 'full')])
        return s_order_lines

    def get_prod_purchase_order_lines(self, prod):
        """
        Returns purchase order line for given product
        :param prod: product object
        :return: Objects of purchase order line for that product
        """
        po_lines = self.env['purchase.order.line'].search(
            [('product_id', '=', prod.id), ('order_id.state', '!=', 'cancel'),
             ('order_id.receipt_status', '!=', 'full')])
        return po_lines

    def auto_email_for_low_stock_product(self):
        """
        Schedular to send email for product which have stock below Threshold point.
        :return: Boolean
        """
        low_stock_data = []
        for prod in self.env['product.product'].search([]):
            if prod.qty_available < prod.qty_available_threshold:
                low_stock_data.append({
                    'product_name': prod.name,
                    'default_code': prod.default_code,
                    'qty_available': prod.qty_available,
                    'min_qty': prod.qty_available_threshold,
                })
        template = self.env.ref('stock_extended_ept.email_template_low_stock_alert')
        if low_stock_data and template and template.email_to:
            company = self.env.company
            if not template.email_from:
                template.email_from = self.env.user.email
            template.with_context(datas=low_stock_data).send_mail(
                company.id,
                force_send=True
            )
        return True
