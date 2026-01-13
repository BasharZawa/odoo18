# -*- coding: utf-8 -*-

import io
import math
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, PatternFill
from datetime import datetime
from odoo import models, fields


class ComponentAvailabilityWizard(models.TransientModel):
    _name = 'component.availability.wizard'
    _description = 'Component Availability Wizard'

    bom_ids = fields.Many2many(comodel_name="mrp.bom", string='Bill of Materials', required=True)

    def generate_report(self):
        """
        Wizard button method to generate attachment for selected BOMs
        :return: action dictionary
        """
        bom_data = self.prepare_selected_bom_stock_data()
        if bom_data:
            attachment = self.create_xlsx_attachment(bom_data)
            return {
                'type': 'ir.actions.act_url',
                'url': f"/web/content/{attachment.id}?download=true",
                'target': 'self'
            }

    def prepare_selected_bom_stock_data(self):
        """
        Prepare BOM details
        :return: bom_data - List of dictionary
        """
        bom_data = {}
        sel_boms = self.bom_ids
        for bom in sel_boms:
            product = bom.product_id or (bom.product_tmpl_id and
                                         bom.product_tmpl_id.product_variant_ids[0])
            if product:
                self.add_current_bom_data(product, bom, bom_data)
                comp_data = self.get_bom_components_data(product, bom, bom_data)
                if comp_data:
                    total_cost = self.calculate_dead_stock_and_cost(comp_data)
                    bom_data[bom].update({
                        'comp_data': comp_data,
                        'total_cost': total_cost
                    })
        return bom_data

    def create_xlsx_attachment(self, bom_datas):
        """
         Creates xlsx attachment
        :param bom_datas: Prepared BOM details which needs to include in report
        :return: attachment new record object
        """
        wb, styles = self.create_workbook()

        for idx, bom in enumerate(bom_datas, start=1):
            ws = self.prepare_sheet(wb, bom, idx)
            self.set_column_dimensions(ws)

            bom_data = bom_datas[bom]
            comp_data = bom_data.get('comp_data', [])
            total_cost = bom_data.get('total_cost', 0)
            bom_summary = {k: v for k, v in bom_data.items() if
                           k not in ['comp_data', 'total_cost']}

            self.write_bom_summary(ws, bom_summary, styles)
            headers = comp_data and list(comp_data[0].keys()) or []
            # headers = headers[:-1] + [''] + headers[-1:]

            self.write_headers(ws, headers, styles)
            self.write_component_rows(ws, comp_data, headers, styles)
            self.add_total_cost(ws, total_cost)

        return self.create_bom_workbook_attachment(wb)

    def create_workbook(self):
        """
        Create workbook and fetch style detail to add in workbook
        :return: Workbook object, Generic Style data
        """
        wb = Workbook()
        styles = self.get_fills_and_borders()
        return wb, styles

    @staticmethod
    def prepare_sheet(wb, bom, idx):
        """
        Prepare sheet inside workbook for specific BOM
        :param wb: Workbook object
        :param bom: Current BOM object
        :param idx: Index of sheet
        :return:
        """
        reference = bom.product_id.default_code or bom.product_tmpl_id.default_code or bom.display_name
        ws = wb.create_sheet(title=f"Sheet{idx}") if idx > 1 else wb.active
        ws.title = reference
        return ws

    @staticmethod
    def set_column_dimensions(ws):
        """
        Add dimensions to columns in sheet
        :param ws: Worksheet object
        :return:
        """
        widths = {
            'A': 30, 'B': 20, 'E': 40, 'F': 10, 'G': 10, 'H': 10, 'I': 11,
            'J': 17, 'K': 10, 'L': 10, 'M': 13, 'N': 13, 'O': 15,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.row_dimensions[7].height = 70

    @staticmethod
    def get_fills_and_borders():
        """
        Prepare object for fills and borders
        :return: fills and borders style details
        """
        fills = {
            'yellow': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            'cyan': PatternFill(start_color="ccccff", end_color="ccccff", fill_type="solid"),
            'green': PatternFill(start_color="b3e6b3", end_color="b3e6b3", fill_type="solid"),
            'violet': PatternFill(start_color="e6b3cc", end_color="e6b3cc", fill_type="solid"),
        }
        thin = Side(border_style="thin", color="000000")
        borders = {
            'thin': Border(left=thin, right=thin, top=thin, bottom=thin)
        }
        return {'fills': fills, 'borders': borders}

    @staticmethod
    def write_bom_summary(ws, bom_data, styles):
        """
        Write down top BOM summary
        :param ws: Current worksheet
        :param bom_data: BOM data dictionary
        :param styles: Style dictionary
        :return: None
        """
        row = 1
        for key, value in bom_data.items():
            ws.cell(row=row, column=1, value=key)
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.fill = styles['fills']['cyan'] if row in [2, 6] else styles['fills']['yellow']
            row += 1

    @staticmethod
    def write_headers(ws, headers, styles):
        """
        Write down all headers for component data
        :param ws: Current worksheet
        :param headers: List of headers
        :param styles: Style dictionary
        :return: None
        """
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.font = Font(bold=True)
            cell.border = styles['borders']['thin']

    @staticmethod
    def write_component_rows(ws, comp_data, headers, styles):
        """
        Fill up data related to components in sheet
        :param ws: Worksheet object
        :param comp_data: Component data dictionary
        :param headers: List of headers
        :param styles: Style data dictionary
        :return: None
        """
        for i, row in enumerate(comp_data, 8):
            for j, col in enumerate(headers, 1):
                cell = ws.cell(row=i, column=j)
                cell.value = row.get(col, "")
                cell.border = styles['borders']['thin']

                if col in ['Qty on\ncurrent P\norder/s', 'min order\nQty\nacceptable\nto supplier',
                           'PO No']:
                    cell.fill = styles['fills']['yellow']
                elif col in ['Total Qty\nafter\nreceiving\nqty on\norder',
                             'Remaining after\nproduction', 'Required', 'Shortage',
                             'Qty to\nbe Ordered']:
                    cell.fill = styles['fills']['cyan']
                elif col in ['Producible\nunits', 'Producible units\nafter new order']:
                    cell.fill = styles['fills']['green']

    @staticmethod
    def add_total_cost(ws, total_cost):
        """
        Add Total cost in a cell value in last column
        :param ws: Worksheet object
        :param total_cost: Total cost
        :return: None
        """
        ws.cell(row=ws.max_row + 2, column=ws.max_column, value=total_cost)

    def create_bom_workbook_attachment(self, wb):
        """
        Save data of workbook in an attachment
        :param wb: Workbook object
        :return: Newly created attachment object
        """
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        file_content = stream.getvalue()
        fmt = '%m-%d-%Y'

        return self.env['ir.attachment'].create({
            'name': f'BOM Components availability({datetime.now().strftime(fmt)}).xlsx',
            'type': 'binary',
            'datas': base64.b64encode(file_content),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

    def add_current_bom_data(self, product, bom, bom_data):
        """
        Add BOM details in main dictionary
        :param product: product object which used in BOM
        :param bom:
        :param bom_data:
        :return:
        """
        bom_total_sales_qty = self.get_total_sales_qty(product)
        bom_total_replenish_qty = self.get_min_stock_level(product)
        min_producible_bom_qty = self.get_min_producible_bom_qty(bom)
        bom_qty_forcast = 0
        surplus_deficit_qty = (product.qty_available + min_producible_bom_qty -
                               bom_total_sales_qty - bom_qty_forcast - bom_total_replenish_qty)
        bom_data.update({
            bom: {
                'Produced units at hand': product.qty_available,
                'Producible units from current stock': min_producible_bom_qty,
                'Qty on order from customers': bom_total_sales_qty,
                'Qty on forecast': bom_qty_forcast,
                'Min stock level': bom_total_replenish_qty,
                'Surplus/(deficit)': surplus_deficit_qty
            }
        })

    def get_bom_components_data(self, product, bom, bom_data):
        """
        Prepare component data
        :param product: product object
        :param bom: BOM object
        :param bom_data: Bom data dictionary
        :return: comp_data: List of dictionary
        """
        comp_data = []
        min_producible_bom_qty = bom_data.get(bom, '').get('Producible units from current stock', 0)
        surplus_deficit_qty = bom_data.get(bom, '').get('Surplus/(deficit)', 0)
        for line in bom.bom_line_ids:
            comp = line.product_id
            qty_data = self.get_comp_qty_data(comp)
            po_lines = qty_data['po_lines']
            producible_units = math.floor(
                (qty_data['on_hand_qty'] + qty_data['po_qty']) / line.product_qty)
            min_supplier_qty = self.get_min_supplier_qty(comp)
            remained_after_prod = qty_data['on_hand_qty'] + qty_data['po_qty'] - (
                        line.product_qty * min_producible_bom_qty)
            required_qty = -surplus_deficit_qty * line.product_qty if surplus_deficit_qty < 0 else 0
            shortage = required_qty - remained_after_prod if required_qty > remained_after_prod else 0
            data = {
                'Finished Good Item Number': product.default_code,
                'Component': comp.default_code,
                'Current\nCost': comp.standard_price,
                'QTY': line.product_qty,
                'Item Description': comp.name,
                'On-hand\nQTY': qty_data['on_hand_qty'],
                'Qty on\ncurrent P\norder/s': qty_data['po_qty'],
                'Total Qty\nafter\nreceiving\nqty on\norder': qty_data['on_hand_qty'] + qty_data[
                    'po_qty'],
                'Producible\nunits': producible_units,
                'Remaining after\nproduction': remained_after_prod,
                'Required': required_qty,
                'Shortage': shortage,
                'min order\nQty\nacceptable\nto supplier': min_supplier_qty,
                'Qty to\nbe Ordered': 0 if not shortage else min([required_qty, shortage]),
                'PO No': ', '.join(list(set(po_lines.mapped('order_id.name')))),
            }
            comp_data.append(data)
        return comp_data

    def get_comp_qty_data(self, comp):
        """
        Prepare qty detail dictionary
        :param comp: comp object
        :return: qty dictionary for components
        """
        on_hand_qty = comp.qty_available
        po_lines = self.get_current_po_qty(comp)
        po_qty = sum([pol.product_qty - pol.qty_received for pol in po_lines])
        return {
            'on_hand_qty': on_hand_qty or 0,
            'po_lines': po_lines,
            'po_qty': po_qty or 0
        }

    def get_current_po_qty(self, prod):
        """
        Return PO lines object
        :param prod: product object
        :return: obj of po lines
        """
        po_lines = self.env['purchase.order.line'].search(
            [('product_id', '=', prod.id), ('order_id.state', '!=', 'cancel'),
             ('order_id.receipt_status', '!=', 'full')])
        return po_lines

    @staticmethod
    def get_min_supplier_qty(comp):
        """
        Fetch min qty from vendors
        :param comp: comp object
        :return: min qty
        """
        sellers = comp.mapped('seller_ids')
        qty = sellers and sellers[0].min_qty or 0
        return qty

    def get_min_producible_bom_qty(self, bom):
        """
        Return producible qty of bom
        :param bom: bom object
        :return: Producible units based on qty
        """
        all_producible_unit = []
        for line in bom.bom_line_ids:
            comp = line.product_id
            qty_data = self.get_comp_qty_data(comp)
            producible_units = math.floor(
                (qty_data['on_hand_qty'] + qty_data['po_qty']) / line.product_qty)
            all_producible_unit.append(producible_units)
        if all_producible_unit:
            return min(all_producible_unit)
        return 0

    def get_total_sales_qty(self, prod):
        """
        Fetch qty of total sales qty
        :param prod: product object
        :return: Total qty which is in sale order
        """
        s_order_lines = self.env['sale.order.line'].search(
            [('product_id', '=', prod.id), ('order_id.state', '!=', 'cancel'),
             ('order_id.delivery_status', '!=', 'full')]
        )
        total_qty = sum([sol.product_uom_qty - sol.qty_delivered for sol in s_order_lines])
        return total_qty

    def get_min_stock_level(self, prod):
        """
        Fetch minimum stock level from Replenishment
        :param prod:
        :return: Return minimum stock level from Replenishment
        """
        # rpl = self.env['stock.warehouse.orderpoint'].search(
        #     [('product_id', '=', prod.id)], limit=1
        # )
        min_stock_level = prod.qty_available_threshold or 0
        return min_stock_level

    @staticmethod
    def calculate_dead_stock_and_cost(comp_data):
        """
        Add data in comp_data and calculates total cost
        :param comp_data: Comp data
        :return: Total cost
        """
        total_cost = 0
        for data in comp_data:
            cost = data.get('Current\nCost', 0) * data.get('Remaining after\nproduction', 0)
            data.update({
                'Cost': cost
            })
            total_cost +=  cost
        return total_cost
